import random
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Font, Color, Side, Alignment
from openpyxl.styles import colors, alignment
import datetime


dt=datetime.date.today()
today=str(dt.month)+'_'+str(dt.day)+'_'+str(dt.year)
wbook=Workbook()    #set workbook
sheet=wbook.active  #create a worksheet and set it equal to default sheet
sheet.title="Randomization"  #name default worksheet
studyno=raw_input("What is the study number?: ")
subjects=int(raw_input("How many subjects are there?: "))   #set subject calculation variables
titlefont = Font(size=18)
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
center_align = Alignment(horizontal='center')

#Header_Footer
sheet.header_footer.center_header.text = str(studyno)
#sheet.HeaderFooter.center_header.font_size = 18
#sheet.HeaderFooter.center_header.text = str(studyno)

sheet.column_dimensions['A'].width = 30.0
sheet['A1']="Randomized Subjects"
sheet['A1'].font = titlefont
sheet['C1']=str(studyno)
sheet['C1'].font = titlefont
sublist=random.sample(range(1,subjects+1),subjects)           #define empty subject list filled with random numbers NO DUPLICATES
start=2
redFill= PatternFill(fill_type='solid', start_color="ff8080")  #sets background color to red for highlighting

def print_Cells(count):  #function that prints random numbers into appropriate cells
    for y in sublist:
        sheet['A'+str(count)]=y
        sheet['A'+str(count)].border = thin_border
        sheet['A'+str(count)].alignment = center_align
        count+=1

def highlightcells(count):  #function that highlights 5% of the first randomized samples. These samples are the reported samples
    percent=round(subjects*0.05)
    while percent>0:
        currentcell=sheet['A'+str(count)]
        currentcell.fill = redFill
        percent-=1
        count+=1
print "randomization created"

def save_workbook():         #function that prompts to save the excel sheet
    saves=raw_input("Would you like to save your randomization? y/n: ")

    if saves=="y" or saves=="Y":
        wbook.save(str(studyno)+'_randomization_for_chromatography_reporting_'+str(today)+'.xlsx')
    else:
        print "!! randomization deleted !!"
#call all functions in order to generate the randomization sheet
print_Cells(start)
highlightcells(start)
save_workbook()
